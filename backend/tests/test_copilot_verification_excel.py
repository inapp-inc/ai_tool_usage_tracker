"""Tests for Copilot calculation verification Excel export."""

from datetime import UTC, datetime
from pathlib import Path

from openpyxl import load_workbook

from app.collector.adapters.copilot_dump import CopilotPullDumper
from app.collector.adapters.copilot_verification_excel import (
    VERIFICATION_FILENAME,
    build_verification_rows,
    write_calculation_verification_excel,
)


def test_build_verification_rows_tracks_api_parse_pull_and_ingest() -> None:
    raw_pages = [
        {
            "source": "copilot-metrics-rows",
            "page": 1,
            "request": {"day": "2026-06-01", "download_url": "https://example.com/r.json"},
            "response": [
                {
                    "user_login": "dev@acme.example",
                    "user_id": 42,
                    "day": "2026-06-01",
                    "used_chat": True,
                    "used_cli": True,
                    "totals_by_cli": {
                        "token_usage": {
                            "prompt_tokens_sum": 1000,
                            "output_tokens_sum": 500,
                        }
                    },
                }
            ],
        },
        {
            "source": "copilot-seats",
            "page": 1,
            "request": {"org": "acme"},
            "response": {
                "seats": [
                    {
                        "assignee": {"login": "dev@acme.example"},
                        "plan_type": "business",
                        "last_activity_at": "2026-06-01T12:00:00Z",
                    }
                ]
            },
        },
    ]

    api_calls, user_rows, seat_rows, final_rows = build_verification_rows(
        raw_pages,
        ingested_vendor_ids={"copilot-user-42-2026-06-01"},
        skipped_duplicate_vendor_ids=set(),
    )
    assert len(api_calls) == 2
    assert len(user_rows) == 1
    assert user_rows[0]["raw_prompt_tokens"] == 1000
    assert user_rows[0]["token_bind_rule"].startswith("totals_by_cli")
    assert user_rows[0]["included_in_final_pull"] == "Y"
    assert user_rows[0]["ingest_status"] == "ingested_new"
    assert len(seat_rows) == 1
    assert seat_rows[0]["included_in_final_pull"] == "N"
    assert seat_rows[0]["pull_skip_reason"] == "metrics row exists for same user+day"
    assert len(final_rows) == 1
    assert final_rows[0]["vendor_event_id"] == "copilot-user-42-2026-06-01"


def test_build_verification_rows_seat_synthetic_when_metrics_missing() -> None:
    raw_pages = [
        {
            "source": "copilot-metrics-rows",
            "page": 1,
            "request": {
                "day": "2026-06-01",
                "download_url": "",
                "download_error": "Metrics API unavailable — enable Copilot usage metrics in GitHub org",
            },
            "response": [],
        },
        {
            "source": "copilot-seats",
            "page": 1,
            "request": {"org": "acme"},
            "response": {
                "seats": [
                    {
                        "assignee": {"login": "alice"},
                        "plan_type": "business",
                        "last_activity_at": "2026-06-01T12:00:00Z",
                    }
                ]
            },
        },
    ]

    _api_calls, user_rows, _seat_rows, final_rows = build_verification_rows(
        raw_pages,
        fallback_at=datetime(2026, 6, 17, tzinfo=UTC),
    )
    assert len(user_rows) == 2
    assert user_rows[0]["source"] == "copilot-metrics-download-failed"
    assert user_rows[0]["download_error"]
    assert user_rows[1]["source"] == "seat-synthetic-fallback"
    assert user_rows[1]["raw_user_login"] == "alice"
    assert len(final_rows) == 1


def test_build_verification_rows_parsed_records_fallback() -> None:
    raw_pages = [
        {
            "source": "copilot-seats",
            "page": 1,
            "request": {"org": "acme"},
            "response": {"seats": []},
        }
    ]
    parsed_records = [
        {
            "vendor_event_id": "copilot-user-9-2026-06-01",
            "user_email": "bob",
            "occurred_at": "2026-06-01T00:00:00+00:00",
            "input_tokens": 12,
            "output_tokens": 3,
            "model": "copilot-chat",
            "estimated_cost": "0",
        }
    ]

    _api_calls, user_rows, _seat_rows, final_rows = build_verification_rows(
        raw_pages,
        parsed_records=parsed_records,
        ingested_vendor_ids={"copilot-user-9-2026-06-01"},
    )
    assert len(user_rows) == 1
    assert user_rows[0]["source"] == "parsed-records-fallback"
    assert user_rows[0]["parsed_input_tokens"] == 12
    assert user_rows[0]["ingest_status"] == "ingested_new"
    assert len(final_rows) == 1


def test_copilot_pull_dumper_writes_full_verification_excel(tmp_path: Path) -> None:
    since = datetime(2026, 6, 1, tzinfo=UTC)
    until = datetime(2026, 6, 17, tzinfo=UTC)
    dumper = CopilotPullDumper(tmp_path, tool_id="tool-copilot", since=since, until=until)

    dumper.write_metrics_rows(
        day="2026-06-01",
        download_url="https://example.com/report.json",
        rows=[
            {
                "user_login": "alice",
                "user_id": 7,
                "day": "2026-06-01",
                "used_chat": True,
                "totals_by_cli": {
                    "token_usage": {"prompt_tokens_sum": 50, "output_tokens_sum": 25}
                },
            }
        ],
    )
    dumper.write_raw_page(
        source="copilot-seats",
        page=1,
        request_body={"org": "acme", "page": 1},
        response_payload={"seats": []},
        status_code=200,
    )
    dumper.set_ingest_audit(
        ingested_vendor_ids=["copilot-user-7-2026-06-01"],
        skipped_duplicate_vendor_ids=[],
    )
    dumper.write_summary(
        pulled=1,
        ingested=1,
        skipped_duplicates=0,
        since=since,
        until=until,
    )

    excel_path = dumper.run_dir / VERIFICATION_FILENAME
    assert excel_path.exists()
    workbook = load_workbook(excel_path)
    assert "Summary" in workbook.sheetnames
    assert "ApiCalls" in workbook.sheetnames
    assert "UserMetrics" in workbook.sheetnames
    assert "Seats" in workbook.sheetnames
    assert "FinalPull" in workbook.sheetnames
    assert "FieldMapping" in workbook.sheetnames

    user_metrics = workbook["UserMetrics"]
    assert user_metrics["A1"].value == "Source"
    headers = [user_metrics.cell(row=1, column=col).value for col in range(1, user_metrics.max_column + 1)]
    assert "Raw API row (JSON)" in headers
    assert "Ingest status" in headers


def test_write_calculation_verification_excel_summary_totals(tmp_path: Path) -> None:
    raw_pages = [
        {
            "source": "copilot-metrics-rows",
            "page": 1,
            "request": {"day": "2026-06-01"},
            "response": [
                {
                    "user_login": "a",
                    "user_id": 1,
                    "day": "2026-06-01",
                    "used_chat": True,
                    "totals_by_cli": {
                        "token_usage": {"prompt_tokens_sum": 10, "output_tokens_sum": 5}
                    },
                },
                {
                    "user_login": "b",
                    "user_id": 2,
                    "day": "2026-06-01",
                    "used_cli": True,
                    "totals_by_cli": {
                        "token_usage": {"prompt_tokens_sum": 20, "output_tokens_sum": 0}
                    },
                },
            ],
        }
    ]
    output_path = tmp_path / VERIFICATION_FILENAME
    write_calculation_verification_excel(
        output_path=output_path,
        raw_pages=raw_pages,
        pulled=2,
        ingested=2,
        skipped_duplicates=0,
    )
    workbook = load_workbook(output_path)
    summary = workbook["Summary"]
    assert summary["B11"].value == 35
