"""Tests for Cursor calculation verification Excel export."""

from datetime import UTC, datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

from app.collector.adapters.cursor_dump import CursorPullDumper
from app.collector.adapters.cursor_verification_excel import (
    VERIFICATION_FILENAME,
    build_verification_rows,
    write_calculation_verification_excel,
)
from app.collector.adapters.usage_parsing import parse_cursor_usage_page


def test_build_verification_rows_flags_included_kind_plan_cost() -> None:
    payload = {
        "usageEvents": [
            {
                "timestamp": "1781804248151",
                "userEmail": "dev@acme.example",
                "model": "default",
                "kind": "Included in Business",
                "tokenUsage": {
                    "inputTokens": 100,
                    "outputTokens": 50,
                    "cacheWriteTokens": 10,
                    "cacheReadTokens": 5,
                },
                "chargedCents": 999,
            }
        ]
    }
    raw_pages = [
        {
            "source": "filtered-usage-events",
            "page": 1,
            "response": payload,
        }
    ]
    filtered_rows, daily_rows = build_verification_rows(raw_pages)
    assert len(filtered_rows) == 1
    assert filtered_rows[0]["kind_included"] == "Y"
    assert filtered_rows[0]["parsed_estimated_cost"] == 0.0
    assert filtered_rows[0]["parsed_included_cost"] == 9.99
    assert filtered_rows[0]["parsed_total_cost"] == 9.99
    assert filtered_rows[0]["parsed_total_tokens"] == 165
    assert filtered_rows[0]["cost_matches_rule"] == "Y"
    assert daily_rows == []


def test_cursor_pull_dumper_writes_verification_excel(tmp_path: Path) -> None:
    since = datetime(2026, 6, 1, tzinfo=UTC)
    until = datetime(2026, 6, 17, tzinfo=UTC)
    dumper = CursorPullDumper(tmp_path, tool_id="tool-123", since=since, until=until)

    payload = {
        "usageEvents": [
            {
                "timestamp": "1781804248151",
                "userEmail": "dev@acme.example",
                "model": "default",
                "kind": "Usage-based",
                "tokenUsage": {"inputTokens": 10, "outputTokens": 5},
                "chargedCents": 12.5,
            }
        ],
        "pagination": {"hasNextPage": False},
    }
    dumper.write_raw_page(
        source="filtered-usage-events",
        page=1,
        request_body={"page": 1},
        response_payload=payload,
        status_code=200,
    )
    records, _ = parse_cursor_usage_page(payload)
    dumper.add_parsed_records(records)
    dumper.write_parsed_records()
    dumper.write_summary(
        pulled=len(records),
        ingested=1,
        skipped_duplicates=0,
        since=since,
        until=until,
    )

    excel_path = dumper.run_dir / VERIFICATION_FILENAME
    assert excel_path.exists()
    workbook = load_workbook(excel_path)
    assert "Summary" in workbook.sheetnames
    assert "FilteredEvents" in workbook.sheetnames
    assert records[0].estimated_cost == Decimal("0.125")


def test_write_calculation_verification_excel_summary_totals(tmp_path: Path) -> None:
    raw_pages = [
        {
            "source": "filtered-usage-events",
            "page": 1,
            "response": {
                "usageEvents": [
                    {
                        "timestamp": "1",
                        "userEmail": "a@example.com",
                        "model": "m1",
                        "kind": "Usage-based",
                        "tokenUsage": {"inputTokens": 10, "outputTokens": 5},
                        "chargedCents": 100,
                    },
                    {
                        "timestamp": "2",
                        "userEmail": "b@example.com",
                        "model": "m2",
                        "kind": "Usage-based",
                        "tokenUsage": {"inputTokens": 20, "outputTokens": 0},
                        "chargedCents": 50,
                    },
                ]
            },
        }
    ]
    output_path = tmp_path / VERIFICATION_FILENAME
    write_calculation_verification_excel(
        output_path=output_path,
        raw_pages=raw_pages,
        pulled=2,
        ingested=2,
        skipped_duplicates=0,
    )
    workbook = load_workbook(output_path)
    summary = workbook["Summary"]
    assert summary["B7"].value == 35
    assert summary["B8"].value == 0
    assert summary["B9"].value == 1.5
