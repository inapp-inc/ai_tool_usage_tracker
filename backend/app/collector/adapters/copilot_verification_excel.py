"""Build Excel workbook: GitHub Copilot API → parse → pull merge → DB ingest."""

from __future__ import annotations

import json
from datetime import UTC, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from app.collector.adapters.base import UsageRecord
from app.collector.adapters.copilot_merge import merge_copilot_records, metrics_overlap_exists
from app.collector.adapters.copilot_metrics_fields import (
    activity_fallback,
    cli_token_sums,
    tokens_from_user_day,
    user_login_from_row,
)
from app.collector.adapters.copilot_parsing import (
    parse_copilot_seat,
    parse_copilot_user_day,
    seat_to_synthetic_user_day,
)
from app.usage.tokens import usage_event_token_total

VERIFICATION_FILENAME = "calculation-verification.xlsx"

API_CALL_COLUMNS = [
    ("source", "Source"),
    ("page", "Page"),
    ("status_code", "HTTP status"),
    ("request_summary", "Request"),
    ("response_summary", "Response summary"),
    ("row_count", "Rows in response"),
]

USER_METRICS_COLUMNS = [
    ("source", "Source"),
    ("api_day", "API report day"),
    ("download_url", "Download URL"),
    ("raw_day", "API day field"),
    ("raw_user_login", "API user_login"),
    ("raw_user_id", "API user_id"),
    ("raw_used_chat", "API used_chat"),
    ("raw_used_cli", "API used_cli"),
    ("raw_used_agent", "API used_agent"),
    ("raw_prompt_tokens", "API CLI prompt_tokens_sum"),
    ("raw_output_tokens", "API CLI output_tokens_sum"),
    ("raw_loc_added", "API loc_added_sum"),
    ("raw_loc_deleted", "API loc_deleted_sum"),
    ("raw_codegen_count", "API code_generation_activity_count"),
    ("raw_accept_count", "API code_acceptance_activity_count"),
    ("raw_interaction_count", "API user_initiated_interaction_count"),
    ("download_error", "Download error"),
    ("token_bind_rule", "Token bind rule"),
    ("parsed_input_tokens", "Parsed input_tokens"),
    ("parsed_output_tokens", "Parsed output_tokens"),
    ("parsed_total_tokens", "Parsed total_tokens"),
    ("parsed_model", "Parsed model"),
    ("parsed_occurred_at", "Parsed occurred_at (UTC)"),
    ("parsed_user_email", "Parsed user_email"),
    ("parsed_estimated_cost", "Parsed estimated_cost (USD)"),
    ("vendor_event_id", "vendor_event_id"),
    ("parse_produced_record", "Parser produced record?"),
    ("included_in_final_pull", "Included in final pull?"),
    ("pull_skip_reason", "Pull skip reason"),
    ("ingest_status", "Ingest status"),
    ("raw_json", "Raw API row (JSON)"),
]

SEAT_COLUMNS = [
    ("source", "Source"),
    ("page", "Page"),
    ("raw_assignee_login", "API assignee.login"),
    ("raw_plan_type", "API plan_type"),
    ("raw_last_activity_at", "API last_activity_at"),
    ("raw_last_activity_editor", "API last_activity_editor"),
    ("raw_created_at", "API created_at"),
    ("raw_updated_at", "API updated_at"),
    ("seat_bind_rule", "Seat bind rule"),
    ("parsed_input_tokens", "Parsed input_tokens"),
    ("parsed_output_tokens", "Parsed output_tokens"),
    ("parsed_total_tokens", "Parsed total_tokens"),
    ("parsed_model", "Parsed model"),
    ("parsed_occurred_at", "Parsed occurred_at (UTC)"),
    ("parsed_user_email", "Parsed user_email"),
    ("vendor_event_id", "vendor_event_id"),
    ("parse_produced_record", "Parser produced record?"),
    ("included_in_final_pull", "Included in final pull?"),
    ("pull_skip_reason", "Pull skip reason"),
    ("ingest_status", "Ingest status"),
    ("raw_json", "Raw API seat (JSON)"),
]

FINAL_PULL_COLUMNS = [
    ("vendor_event_id", "vendor_event_id"),
    ("record_source", "Record source"),
    ("user_email", "user_email"),
    ("user_name", "user_name"),
    ("model", "model"),
    ("occurred_at", "occurred_at (UTC)"),
    ("input_tokens", "input_tokens"),
    ("output_tokens", "output_tokens"),
    ("total_tokens", "total_tokens"),
    ("estimated_cost", "estimated_cost (USD)"),
    ("requests", "requests"),
    ("ingest_status", "Ingest status"),
    ("db_provider", "DB provider"),
    ("db_bind_note", "DB column bind note"),
]

FIELD_MAPPING_ROWS = [
    ("User metrics", "totals_by_cli.token_usage.prompt_tokens_sum", "input_tokens", "Primary token source when > 0"),
    ("User metrics", "totals_by_cli.token_usage.output_tokens_sum", "output_tokens", "Primary token source when > 0"),
    ("User metrics", "totals_by_ide/feature breakdown arrays", "input/output tokens", "Second priority when CLI tokens are 0"),
    ("User metrics", "loc_added_sum + activity counts", "input_tokens", "Fallback when breakdown arrays are empty"),
    ("User metrics", "used_* flags (chat/cli/agent/cloud/review)", "model + zero-token record", "Activity with no token fields"),
    ("User metrics", "day / report_day", "occurred_at", "UTC midnight on report day"),
    ("User metrics", "user_login / login", "user_email / user_name", "GitHub username"),
    ("User metrics", "user_id + day", "vendor_event_id", "copilot-user-{user_id}-{day}"),
    ("Seats fallback", "assignee.login + last_activity_at", "synthetic user row", "Used when metrics NDJSON unavailable"),
    ("Seats", "last_activity_at present", "input_tokens=1", "Seat activity marker"),
    ("Seats", "assignee.login", "user_email / user_name", "GitHub login"),
    ("Merge", "metrics row exists for user+day", "seat row skipped", "Metrics preferred over seat fallback"),
    ("Ingest", "vendor_event_id", "usage_events.vendor_event_id", "Unique per provider"),
]


def _token_bind_rule(row: dict[str, Any], prompt: int, output: int, fallback: int) -> str:
    _input, _output, rule = tokens_from_user_day(row)
    return rule


def _build_user_metric_row(
    row: dict[str, Any],
    *,
    source: str,
    api_day: str,
    download_url: str,
    download_error: str | None,
    final_ids: set[str],
    interim_metrics: list,
    ingested_vendor_ids: set[str],
    skipped_duplicate_vendor_ids: set[str],
) -> dict[str, Any]:
    prompt, output = cli_token_sums(row)
    fallback = activity_fallback(row)
    parsed = parse_copilot_user_day(row)
    vendor_event_id = parsed.vendor_event_id if parsed else None
    in_final = vendor_event_id in final_ids if vendor_event_id else False
    pull_skip = ""
    if parsed and not in_final:
        pull_skip = "duplicate vendor_event_id in metrics"
    elif parsed is None:
        pull_skip = download_error or "parser returned None"

    if parsed is not None and parsed.vendor_event_id not in {
        getattr(record, "vendor_event_id", None) for record in interim_metrics
    }:
        interim_metrics.append(parsed)

    return {
        "source": source,
        "api_day": api_day,
        "download_url": download_url,
        "download_error": download_error,
        "raw_day": row.get("day") or row.get("report_day"),
        "raw_user_login": user_login_from_row(row),
        "raw_user_id": row.get("user_id"),
        "raw_used_chat": row.get("used_chat"),
        "raw_used_cli": row.get("used_cli"),
        "raw_used_agent": row.get("used_agent"),
        "raw_prompt_tokens": prompt,
        "raw_output_tokens": output,
        "raw_loc_added": row.get("loc_added_sum") or row.get("loc_suggested_to_add_sum"),
        "raw_loc_deleted": row.get("loc_deleted_sum") or row.get("loc_suggested_to_delete_sum"),
        "raw_codegen_count": row.get("code_generation_activity_count"),
        "raw_accept_count": row.get("code_acceptance_activity_count"),
        "raw_interaction_count": row.get("user_initiated_interaction_count"),
        "token_bind_rule": _token_bind_rule(row, prompt, output, fallback),
        "parsed_input_tokens": parsed.input_tokens if parsed else 0,
        "parsed_output_tokens": parsed.output_tokens if parsed else 0,
        "parsed_total_tokens": usage_event_token_total(
            input_tokens=parsed.input_tokens if parsed else 0,
            output_tokens=parsed.output_tokens if parsed else 0,
        ),
        "parsed_model": parsed.model if parsed else None,
        "parsed_occurred_at": parsed.occurred_at.isoformat() if parsed else None,
        "parsed_user_email": parsed.user_email if parsed else None,
        "parsed_estimated_cost": float(parsed.estimated_cost) if parsed else 0.0,
        "vendor_event_id": vendor_event_id,
        "parse_produced_record": "Y" if parsed else "N",
        "included_in_final_pull": "Y" if in_final else "N",
        "pull_skip_reason": pull_skip if not in_final else "",
        "ingest_status": _ingest_status(
            vendor_event_id,
            in_final_pull=in_final,
            ingested_vendor_ids=ingested_vendor_ids,
            skipped_duplicate_vendor_ids=skipped_duplicate_vendor_ids,
        ),
        "raw_json": _raw_json(row),
    }


def _rows_from_parsed_records(
    parsed_records: list[dict[str, Any]],
    *,
    final_ids: set[str],
    ingested_vendor_ids: set[str],
    skipped_duplicate_vendor_ids: set[str],
    existing_vendor_ids: set[str],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for record in parsed_records:
        vendor_event_id = record.get("vendor_event_id")
        if not isinstance(vendor_event_id, str) or not vendor_event_id.startswith("copilot-user-"):
            continue
        if vendor_event_id in existing_vendor_ids:
            continue
        in_final = True
        rows.append(
            {
                "source": "parsed-records-fallback",
                "api_day": str(record.get("occurred_at", ""))[:10],
                "download_url": "",
                "download_error": None,
                "raw_day": str(record.get("occurred_at", ""))[:10],
                "raw_user_login": record.get("user_email"),
                "raw_user_id": None,
                "raw_used_chat": None,
                "raw_used_cli": None,
                "raw_used_agent": None,
                "raw_prompt_tokens": None,
                "raw_output_tokens": None,
                "raw_loc_added": None,
                "raw_loc_deleted": None,
                "raw_codegen_count": None,
                "raw_accept_count": None,
                "raw_interaction_count": None,
                "token_bind_rule": "from parsed-records.json (sync output)",
                "parsed_input_tokens": record.get("input_tokens"),
                "parsed_output_tokens": record.get("output_tokens"),
                "parsed_total_tokens": usage_event_token_total(
                    input_tokens=int(record.get("input_tokens") or 0),
                    output_tokens=int(record.get("output_tokens") or 0),
                ),
                "parsed_model": record.get("model"),
                "parsed_occurred_at": record.get("occurred_at"),
                "parsed_user_email": record.get("user_email"),
                "parsed_estimated_cost": float(record.get("estimated_cost") or 0),
                "vendor_event_id": vendor_event_id,
                "parse_produced_record": "Y",
                "included_in_final_pull": "Y" if in_final else "N",
                "pull_skip_reason": "",
                "ingest_status": _ingest_status(
                    vendor_event_id,
                    in_final_pull=in_final,
                    ingested_vendor_ids=ingested_vendor_ids,
                    skipped_duplicate_vendor_ids=skipped_duplicate_vendor_ids,
                ),
                "raw_json": _raw_json(record),
            }
        )
    return rows


def _usage_records_from_parsed(parsed_records: list[dict[str, Any]]) -> list[UsageRecord]:
    records: list[UsageRecord] = []
    for row in parsed_records:
        vendor_event_id = row.get("vendor_event_id")
        if not isinstance(vendor_event_id, str) or not vendor_event_id:
            continue
        occurred_raw = row.get("occurred_at")
        if isinstance(occurred_raw, str):
            try:
                occurred_at = datetime.fromisoformat(occurred_raw.replace("Z", "+00:00"))
            except ValueError:
                continue
        elif isinstance(occurred_raw, datetime):
            occurred_at = occurred_raw
        else:
            continue
        cost_raw = row.get("estimated_cost")
        try:
            estimated_cost = Decimal(str(cost_raw or 0))
        except Exception:  # noqa: BLE001
            estimated_cost = Decimal(0)
        records.append(
            UsageRecord(
                vendor_event_id=vendor_event_id,
                model=row.get("model") if isinstance(row.get("model"), str) else None,
                occurred_at=occurred_at,
                input_tokens=int(row.get("input_tokens") or 0),
                output_tokens=int(row.get("output_tokens") or 0),
                estimated_cost=estimated_cost,
                user_email=row.get("user_email") if isinstance(row.get("user_email"), str) else None,
                user_name=row.get("user_name") if isinstance(row.get("user_name"), str) else None,
                requests=int(row.get("requests") or 0),
            )
        )
    return records


def _ingest_status(
    vendor_event_id: str | None,
    *,
    in_final_pull: bool,
    ingested_vendor_ids: set[str],
    skipped_duplicate_vendor_ids: set[str],
) -> str:
    if not vendor_event_id or not in_final_pull:
        return "not_in_pull"
    if vendor_event_id in ingested_vendor_ids:
        return "ingested_new"
    if vendor_event_id in skipped_duplicate_vendor_ids:
        return "skipped_duplicate"
    return "not_ingested"


def _raw_json(value: dict[str, Any], limit: int = 4000) -> str:
    text = json.dumps(value, default=str, ensure_ascii=False)
    if len(text) <= limit:
        return text
    return text[: limit - 3] + "..."


def _collect_raw_pages(raw_pages: list[dict[str, Any]]) -> tuple[
    list[dict[str, Any]],
    list[tuple[dict[str, Any], str, str, str | None]],
    list[tuple[dict[str, Any], int]],
    list[tuple[str, str, str | None]],
]:
    api_calls: list[dict[str, Any]] = []
    metric_rows: list[tuple[dict[str, Any], str, str, str | None]] = []
    seat_rows: list[tuple[dict[str, Any], int]] = []
    failed_downloads: list[tuple[str, str, str | None]] = []

    for entry in raw_pages:
        source = str(entry.get("source") or "")
        page = int(entry.get("page") or 0)
        status_code = entry.get("status_code")
        request = entry.get("request") if isinstance(entry.get("request"), dict) else {}
        payload = entry.get("response")
        download_error = request.get("download_error")
        if not isinstance(download_error, str) or not download_error.strip():
            download_error = None

        row_count = 0
        response_summary = ""
        if source == "copilot-seats" and isinstance(payload, dict):
            seats = payload.get("seats")
            row_count = len(seats) if isinstance(seats, list) else 0
            response_summary = f"seats={row_count}"
        elif source == "copilot-metrics-rows" and isinstance(payload, list):
            row_count = len(payload)
            response_summary = f"downloaded_rows={row_count}"
        elif source == "copilot-user-metrics" and isinstance(payload, dict):
            links = payload.get("download_links")
            link_count = len(links) if isinstance(links, list) else 0
            response_summary = f"download_links={link_count}"
        else:
            response_summary = type(payload).__name__

        api_calls.append(
            {
                "source": source,
                "page": page,
                "status_code": status_code,
                "request_summary": json.dumps(request, default=str, ensure_ascii=False)[:500],
                "response_summary": response_summary,
                "row_count": row_count,
            }
        )

        if source == "copilot-metrics-rows" and isinstance(payload, list):
            day = str(request.get("day") or "")
            download_url = str(request.get("download_url") or "")
            if payload:
                for row in payload:
                    if isinstance(row, dict):
                        metric_rows.append((row, day, download_url, download_error))
            elif download_error:
                failed_downloads.append((day, download_url, download_error))
            continue

        if source == "copilot-user-metrics" and isinstance(payload, dict):
            day = str(request.get("day") or "")
            embedded = payload.get("rows")
            if isinstance(embedded, list):
                for row in embedded:
                    if isinstance(row, dict):
                        metric_rows.append((row, day, "", None))
            continue

        if source == "copilot-seats" and isinstance(payload, dict):
            seats = payload.get("seats")
            if isinstance(seats, list):
                for seat in seats:
                    if isinstance(seat, dict):
                        seat_rows.append((seat, page))

    return api_calls, metric_rows, seat_rows, failed_downloads


def build_verification_rows(
    raw_pages: list[dict[str, Any]],
    *,
    parsed_records: list[dict[str, Any]] | None = None,
    ingested_vendor_ids: set[str] | None = None,
    skipped_duplicate_vendor_ids: set[str] | None = None,
    fallback_at: datetime | None = None,
) -> tuple[
    list[dict[str, Any]],
    list[dict[str, Any]],
    list[dict[str, Any]],
    list[dict[str, Any]],
]:
    ingested_vendor_ids = ingested_vendor_ids or set()
    skipped_duplicate_vendor_ids = skipped_duplicate_vendor_ids or set()
    fallback_at = fallback_at or datetime.now(UTC)
    parsed_records = parsed_records or []

    api_calls, metric_rows, seat_payloads, failed_downloads = _collect_raw_pages(raw_pages)

    metrics_records: list[UsageRecord] = []
    for row, _day, _url, _error in metric_rows:
        parsed = parse_copilot_user_day(row)
        if parsed is not None:
            metrics_records.append(parsed)

    final_records = merge_copilot_records(
        metrics_records,
        [seat for seat, _page in seat_payloads],
        fallback_at=fallback_at,
    )
    final_ids = {record.vendor_event_id for record in final_records if record.vendor_event_id}

    interim_metrics: list[UsageRecord] = []
    user_rows: list[dict[str, Any]] = []
    for row, api_day, download_url, download_error in metric_rows:
        user_rows.append(
            _build_user_metric_row(
                row,
                source="copilot-metrics-rows",
                api_day=api_day,
                download_url=download_url,
                download_error=download_error,
                final_ids=final_ids,
                interim_metrics=interim_metrics,
                ingested_vendor_ids=ingested_vendor_ids,
                skipped_duplicate_vendor_ids=skipped_duplicate_vendor_ids,
            )
        )

    for api_day, download_url, download_error in failed_downloads:
        user_rows.append(
            _build_user_metric_row(
                {},
                source="copilot-metrics-download-failed",
                api_day=api_day,
                download_url=download_url,
                download_error=download_error,
                final_ids=final_ids,
                interim_metrics=interim_metrics,
                ingested_vendor_ids=ingested_vendor_ids,
                skipped_duplicate_vendor_ids=skipped_duplicate_vendor_ids,
            )
        )

    if not metric_rows:
        parsed_user_records = [
            record
            for record in parsed_records
            if str(record.get("vendor_event_id") or "").startswith("copilot-user-")
        ]
        if parsed_user_records:
            existing_vendor_ids = {
                vendor_id
                for vendor_id in (row.get("vendor_event_id") for row in user_rows)
                if isinstance(vendor_id, str) and vendor_id
            }
            user_rows.extend(
                _rows_from_parsed_records(
                    parsed_user_records,
                    final_ids=final_ids,
                    ingested_vendor_ids=ingested_vendor_ids,
                    skipped_duplicate_vendor_ids=skipped_duplicate_vendor_ids,
                    existing_vendor_ids=existing_vendor_ids,
                )
            )
        elif seat_payloads:
            fallback_day = fallback_at.date().isoformat()
            for seat, _page in seat_payloads:
                synthetic = seat_to_synthetic_user_day(seat, fallback_day=fallback_day)
                user_rows.append(
                    _build_user_metric_row(
                        synthetic,
                        source="seat-synthetic-fallback",
                        api_day=fallback_day,
                        download_url="",
                        download_error="Metrics NDJSON unavailable — row derived from seat API",
                        final_ids=final_ids,
                        interim_metrics=interim_metrics,
                        ingested_vendor_ids=ingested_vendor_ids,
                        skipped_duplicate_vendor_ids=skipped_duplicate_vendor_ids,
                    )
                )

    interim_after_metrics = list(interim_metrics)
    seat_rows_out: list[dict[str, Any]] = []
    for seat, page in seat_payloads:
        assignee = seat.get("assignee") if isinstance(seat.get("assignee"), dict) else {}
        parsed = parse_copilot_seat(seat, fallback_at=fallback_at)
        vendor_event_id = parsed.vendor_event_id if parsed else None
        in_final = vendor_event_id in final_ids if vendor_event_id else False
        pull_skip = ""
        if parsed is None:
            pull_skip = "missing assignee.login"
        elif not in_final:
            if metrics_overlap_exists(
                interim_after_metrics,
                user_email=parsed.user_email,
                occurred_date=parsed.occurred_at.date(),
            ):
                pull_skip = "metrics row exists for same user+day"
            elif vendor_event_id and vendor_event_id in {
                record.vendor_event_id for record in interim_after_metrics
            }:
                pull_skip = "duplicate vendor_event_id"
            else:
                pull_skip = "excluded by merge"

        seat_rows_out.append(
            {
                "source": "copilot-seats",
                "page": page,
                "raw_assignee_login": assignee.get("login"),
                "raw_plan_type": seat.get("plan_type"),
                "raw_last_activity_at": seat.get("last_activity_at"),
                "raw_last_activity_editor": seat.get("last_activity_editor"),
                "raw_created_at": seat.get("created_at"),
                "raw_updated_at": seat.get("updated_at"),
                "seat_bind_rule": (
                    "last_activity_at → input_tokens=1"
                    if seat.get("last_activity_at")
                    else "no last_activity_at → input_tokens=0"
                ),
                "parsed_input_tokens": parsed.input_tokens if parsed else 0,
                "parsed_output_tokens": parsed.output_tokens if parsed else 0,
                "parsed_total_tokens": usage_event_token_total(
                    input_tokens=parsed.input_tokens if parsed else 0,
                    output_tokens=parsed.output_tokens if parsed else 0,
                ),
                "parsed_model": parsed.model if parsed else None,
                "parsed_occurred_at": parsed.occurred_at.isoformat() if parsed else None,
                "parsed_user_email": parsed.user_email if parsed else None,
                "vendor_event_id": vendor_event_id,
                "parse_produced_record": "Y" if parsed else "N",
                "included_in_final_pull": "Y" if in_final else "N",
                "pull_skip_reason": pull_skip if not in_final else "",
                "ingest_status": _ingest_status(
                    vendor_event_id,
                    in_final_pull=in_final,
                    ingested_vendor_ids=ingested_vendor_ids,
                    skipped_duplicate_vendor_ids=skipped_duplicate_vendor_ids,
                ),
                "raw_json": _raw_json(seat),
            }
        )

    final_pull_rows: list[dict[str, Any]] = []
    pull_records = final_records
    if not pull_records and parsed_records:
        pull_records = _usage_records_from_parsed(parsed_records)

    for record in pull_records:
        vendor_event_id = record.vendor_event_id or ""
        source = "user_metrics" if vendor_event_id.startswith("copilot-user-") else "seat_fallback"
        in_final = True
        final_pull_rows.append(
            {
                "vendor_event_id": vendor_event_id,
                "record_source": source,
                "user_email": record.user_email,
                "user_name": record.user_name,
                "model": record.model,
                "occurred_at": record.occurred_at.isoformat(),
                "input_tokens": record.input_tokens,
                "output_tokens": record.output_tokens,
                "total_tokens": record.total_tokens,
                "estimated_cost": float(record.estimated_cost),
                "requests": getattr(record, "requests", 0),
                "ingest_status": _ingest_status(
                    vendor_event_id,
                    in_final_pull=in_final,
                    ingested_vendor_ids=ingested_vendor_ids,
                    skipped_duplicate_vendor_ids=skipped_duplicate_vendor_ids,
                ),
                "db_provider": "copilot",
                "db_bind_note": "insert into usage.usage_events on sync",
            }
        )

    return api_calls, user_rows, seat_rows_out, final_pull_rows


def _write_sheet(
    workbook: Workbook,
    *,
    title: str,
    columns: list[tuple[str, str]],
    rows: list[dict[str, Any]],
) -> None:
    sheet = workbook.create_sheet(title=title)
    header_font = Font(bold=True)
    keys = [key for key, _label in columns]
    labels = [label for _key, label in columns]

    sheet.append(labels)
    for cell in sheet[1]:
        cell.font = header_font

    for row in rows:
        sheet.append([row.get(key) for key in keys])

    for index, (_key, label) in enumerate(columns, start=1):
        width = min(max(len(label) + 2, 12), 48)
        sheet.column_dimensions[get_column_letter(index)].width = width


def _write_field_mapping_sheet(workbook: Workbook) -> None:
    sheet = workbook.create_sheet(title="FieldMapping")
    header_font = Font(bold=True)
    sheet.append(["Area", "API field(s)", "App field", "Rule"])
    for cell in sheet[1]:
        cell.font = header_font
    for row in FIELD_MAPPING_ROWS:
        sheet.append(list(row))
    sheet.column_dimensions["A"].width = 16
    sheet.column_dimensions["B"].width = 42
    sheet.column_dimensions["C"].width = 28
    sheet.column_dimensions["D"].width = 48


def _write_summary_sheet(
    workbook: Workbook,
    *,
    api_calls: list[dict[str, Any]],
    user_rows: list[dict[str, Any]],
    seat_rows: list[dict[str, Any]],
    final_pull_rows: list[dict[str, Any]],
    pulled: int,
    ingested: int,
    skipped_duplicates: int,
) -> None:
    sheet = workbook.active
    sheet.title = "Summary"
    bold = Font(bold=True)

    sheet["A1"] = "Metric"
    sheet["B1"] = "Value"
    sheet["A1"].font = bold
    sheet["B1"].font = bold

    user_tokens = sum(int(row.get("parsed_total_tokens") or 0) for row in user_rows)
    final_tokens = sum(int(row.get("total_tokens") or 0) for row in final_pull_rows)
    ingested_rows = sum(1 for row in final_pull_rows if row.get("ingest_status") == "ingested_new")
    duplicate_rows = sum(
        1 for row in final_pull_rows if row.get("ingest_status") == "skipped_duplicate"
    )

    metrics = [
        ("API calls captured", len(api_calls)),
        ("User metrics API rows", len(user_rows)),
        ("Seat API rows", len(seat_rows)),
        ("Final pull records", len(final_pull_rows)),
        ("Parsed records pulled (sync)", pulled),
        ("New records ingested (sync)", ingested),
        ("Skipped duplicates (sync)", skipped_duplicates),
        ("Final pull marked ingested_new", ingested_rows),
        ("Final pull marked skipped_duplicate", duplicate_rows),
        ("Sum tokens (user metrics sheet)", user_tokens),
        ("Sum tokens (final pull sheet)", final_tokens),
    ]
    for index, (label, value) in enumerate(metrics, start=2):
        sheet[f"A{index}"] = label
        sheet[f"B{index}"] = value

    sheet.column_dimensions["A"].width = 42
    sheet.column_dimensions["B"].width = 24


def write_calculation_verification_excel(
    *,
    output_path: Path,
    raw_pages: list[dict[str, Any]],
    pulled: int,
    ingested: int,
    skipped_duplicates: int,
    parsed_records: list[dict[str, Any]] | None = None,
    ingested_vendor_ids: set[str] | None = None,
    skipped_duplicate_vendor_ids: set[str] | None = None,
    since: str | None = None,
    until: str | None = None,
) -> Path:
    fallback_at = datetime.now(UTC)
    if until:
        try:
            fallback_at = datetime.fromisoformat(until.replace("Z", "+00:00"))
        except ValueError:
            pass

    api_calls, user_rows, seat_rows, final_pull_rows = build_verification_rows(
        raw_pages,
        parsed_records=parsed_records,
        ingested_vendor_ids=ingested_vendor_ids,
        skipped_duplicate_vendor_ids=skipped_duplicate_vendor_ids,
        fallback_at=fallback_at,
    )

    workbook = Workbook()
    _write_summary_sheet(
        workbook,
        api_calls=api_calls,
        user_rows=user_rows,
        seat_rows=seat_rows,
        final_pull_rows=final_pull_rows,
        pulled=pulled,
        ingested=ingested,
        skipped_duplicates=skipped_duplicates,
    )
    _write_sheet(workbook, title="ApiCalls", columns=API_CALL_COLUMNS, rows=api_calls)
    _write_sheet(workbook, title="UserMetrics", columns=USER_METRICS_COLUMNS, rows=user_rows)
    _write_sheet(workbook, title="Seats", columns=SEAT_COLUMNS, rows=seat_rows)
    _write_sheet(workbook, title="FinalPull", columns=FINAL_PULL_COLUMNS, rows=final_pull_rows)
    _write_field_mapping_sheet(workbook)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path
