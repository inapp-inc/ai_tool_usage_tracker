"""Build a single Excel workbook comparing raw Cursor API fields to parsed values."""

from __future__ import annotations

from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from app.collector.adapters.usage_parsing import (
    _cursor_estimated_cost,
    _cursor_kind_is_included,
    _epoch_ms_to_datetime,
    _int_token,
    cursor_vendor_event_id,
)
from app.usage.tokens import usage_event_token_total

VERIFICATION_FILENAME = "calculation-verification.xlsx"

FILTERED_COLUMNS = [
    ("source", "Source"),
    ("page", "Page"),
    ("timestamp_raw", "Raw timestamp (ms)"),
    ("occurred_at", "Parsed occurred_at (UTC)"),
    ("user_email", "User email"),
    ("model", "Model"),
    ("kind", "Kind"),
    ("kind_included", "Kind included in plan?"),
    ("raw_input_tokens", "API inputTokens"),
    ("raw_output_tokens", "API outputTokens"),
    ("raw_cache_write_tokens", "API cacheWriteTokens"),
    ("raw_cache_read_tokens", "API cacheReadTokens"),
    ("parsed_input_tokens", "Parsed input_tokens"),
    ("parsed_output_tokens", "Parsed output_tokens"),
    ("parsed_cache_write_tokens", "Parsed cache_write_tokens"),
    ("parsed_cache_read_tokens", "Parsed cache_read_tokens"),
    ("parsed_total_tokens", "Parsed total_tokens"),
    ("raw_charged_cents", "API chargedCents"),
    ("raw_charged_usd", "chargedCents / 100"),
    ("cost_rule", "Cost rule applied"),
    ("parsed_estimated_cost", "Parsed estimated_cost (USD)"),
    ("cost_matches_rule", "Cost matches rule?"),
    ("vendor_event_id", "vendor_event_id"),
]

DAILY_COLUMNS = [
    ("source", "Source"),
    ("page", "Page"),
    ("day", "Day"),
    ("email", "Email"),
    ("most_used_model", "Most used model"),
    ("request_total", "Computed request total"),
    ("line_total", "Computed line total"),
    ("parsed_input_tokens", "Parsed input_tokens"),
    ("parsed_output_tokens", "Parsed output_tokens"),
    ("parsed_estimated_cost", "Parsed estimated_cost (USD)"),
    ("vendor_event_id", "vendor_event_id"),
]


def _filtered_row(source: str, page: int, event: dict[str, Any]) -> dict[str, Any]:
    token_usage = event.get("tokenUsage") if isinstance(event.get("tokenUsage"), dict) else {}
    input_tokens = _int_token(token_usage.get("inputTokens"))
    output_tokens = _int_token(token_usage.get("outputTokens"))
    cache_write_tokens = _int_token(token_usage.get("cacheWriteTokens"))
    cache_read_tokens = _int_token(token_usage.get("cacheReadTokens"))
    total_tokens = usage_event_token_total(
        input_tokens=input_tokens,
        output_tokens=output_tokens,
        cache_write_tokens=cache_write_tokens,
        cache_read_tokens=cache_read_tokens,
    )
    kind = event.get("kind")
    kind_included = _cursor_kind_is_included(kind)
    estimated_cost = _cursor_estimated_cost(event, token_usage)
    charged_cents = event.get("chargedCents")
    try:
        raw_charged_usd = Decimal(str(charged_cents or 0)) / Decimal("100")
    except Exception:  # noqa: BLE001
        raw_charged_usd = Decimal("0")

    if kind_included:
        cost_rule = 'kind contains "include" -> $0'
        cost_matches = estimated_cost == Decimal("0")
    else:
        cost_rule = "chargedCents / 100"
        cost_matches = estimated_cost == raw_charged_usd

    timestamp = event.get("timestamp")
    user_email = event.get("userEmail") or event.get("email") or ""
    model = event.get("model") or "cursor-default"

    return {
        "source": source,
        "page": page,
        "timestamp_raw": timestamp,
        "occurred_at": _epoch_ms_to_datetime(timestamp).isoformat(),
        "user_email": user_email,
        "model": model,
        "kind": kind,
        "kind_included": "Y" if kind_included else "N",
        "raw_input_tokens": token_usage.get("inputTokens"),
        "raw_output_tokens": token_usage.get("outputTokens"),
        "raw_cache_write_tokens": token_usage.get("cacheWriteTokens"),
        "raw_cache_read_tokens": token_usage.get("cacheReadTokens"),
        "parsed_input_tokens": input_tokens,
        "parsed_output_tokens": output_tokens,
        "parsed_cache_write_tokens": cache_write_tokens,
        "parsed_cache_read_tokens": cache_read_tokens,
        "parsed_total_tokens": total_tokens,
        "raw_charged_cents": charged_cents,
        "raw_charged_usd": float(raw_charged_usd),
        "cost_rule": cost_rule,
        "parsed_estimated_cost": float(estimated_cost),
        "cost_matches_rule": "Y" if cost_matches else "N",
        "vendor_event_id": cursor_vendor_event_id(
            timestamp=timestamp,
            user_email=user_email,
            model=model,
        ),
    }


def _daily_request_total(row: dict[str, Any]) -> int:
    keys = (
        "chatRequests",
        "composerRequests",
        "agentRequests",
        "cmdkUsages",
        "bugbotUsages",
        "usageBasedReqs",
        "subscriptionIncludedReqs",
        "apiKeyReqs",
    )
    total = 0
    for key in keys:
        try:
            total += int(row.get(key) or 0)
        except (TypeError, ValueError):
            pass
    return total


def _daily_line_total(row: dict[str, Any]) -> int:
    try:
        added = int(row.get("totalLinesAdded") or 0)
    except (TypeError, ValueError):
        added = 0
    try:
        deleted = int(row.get("totalLinesDeleted") or 0)
    except (TypeError, ValueError):
        deleted = 0
    return added + deleted


def _daily_row(source: str, page: int, index: int, row: dict[str, Any]) -> dict[str, Any]:
    request_total = _daily_request_total(row)
    line_total = _daily_line_total(row)
    input_tokens = request_total or line_total
    day = row.get("day")
    email = row.get("email") or ""
    model = row.get("mostUsedModel") or "cursor-daily"
    user_id = row.get("userId", index)
    return {
        "source": source,
        "page": page,
        "day": day,
        "email": email,
        "most_used_model": model,
        "request_total": request_total,
        "line_total": line_total,
        "parsed_input_tokens": max(input_tokens, 0),
        "parsed_output_tokens": 0,
        "parsed_estimated_cost": 0.0,
        "vendor_event_id": f"cursor-daily-{day}-{user_id}-{email}-{index}",
    }


def build_verification_rows(
    raw_pages: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    filtered_rows: list[dict[str, Any]] = []
    daily_rows: list[dict[str, Any]] = []

    for entry in raw_pages:
        source = str(entry.get("source") or "")
        page = int(entry.get("page") or 0)
        payload = entry.get("response")
        if not isinstance(payload, dict):
            continue

        if source == "filtered-usage-events":
            for event in payload.get("usageEvents", []):
                if isinstance(event, dict):
                    filtered_rows.append(_filtered_row(source, page, event))
            continue

        if source == "daily-usage-data":
            for index, row in enumerate(payload.get("data", [])):
                if isinstance(row, dict):
                    daily_rows.append(_daily_row(source, page, index, row))

    return filtered_rows, daily_rows


def _write_sheet(
    workbook: Workbook,
    *,
    title: str,
    columns: list[tuple[str, str]],
    rows: list[dict[str, Any]],
) -> None:
    sheet = workbook.create_sheet(title=title)
    header_font = Font(bold=True)
    keys = [key for key, _label in columns]
    labels = [label for _key, label in columns]

    sheet.append(labels)
    for cell in sheet[1]:
        cell.font = header_font

    for row in rows:
        sheet.append([row.get(key) for key in keys])

    for index, (_key, label) in enumerate(columns, start=1):
        width = min(max(len(label) + 2, 12), 40)
        sheet.column_dimensions[get_column_letter(index)].width = width


def _write_summary_sheet(
    workbook: Workbook,
    *,
    filtered_rows: list[dict[str, Any]],
    daily_rows: list[dict[str, Any]],
    pulled: int,
    ingested: int,
    skipped_duplicates: int,
) -> None:
    sheet = workbook.active
    sheet.title = "Summary"
    bold = Font(bold=True)

    sheet["A1"] = "Metric"
    sheet["B1"] = "Value"
    sheet["A1"].font = bold
    sheet["B1"].font = bold

    filtered_cost = sum(float(row.get("parsed_estimated_cost") or 0) for row in filtered_rows)
    filtered_tokens = sum(int(row.get("parsed_total_tokens") or 0) for row in filtered_rows)
    cost_mismatches = sum(
        1 for row in filtered_rows if row.get("cost_matches_rule") != "Y"
    )
    daily_tokens = sum(int(row.get("parsed_input_tokens") or 0) for row in daily_rows)

    metrics = [
        ("Filtered usage event rows", len(filtered_rows)),
        ("Daily usage rows", len(daily_rows)),
        ("Parsed records pulled", pulled),
        ("New records ingested", ingested),
        ("Skipped duplicates", skipped_duplicates),
        ("Sum parsed total_tokens (filtered)", filtered_tokens),
        ("Sum parsed input_tokens (daily)", daily_tokens),
        ("Sum parsed estimated_cost USD (filtered)", round(filtered_cost, 4)),
        ("Filtered rows with cost rule mismatch", cost_mismatches),
    ]
    for index, (label, value) in enumerate(metrics, start=2):
        sheet[f"A{index}"] = label
        sheet[f"B{index}"] = value

    sheet.column_dimensions["A"].width = 42
    sheet.column_dimensions["B"].width = 24


def write_calculation_verification_excel(
    *,
    output_path: Path,
    raw_pages: list[dict[str, Any]],
    pulled: int,
    ingested: int,
    skipped_duplicates: int,
) -> Path:
    filtered_rows, daily_rows = build_verification_rows(raw_pages)
    workbook = Workbook()
    _write_summary_sheet(
        workbook,
        filtered_rows=filtered_rows,
        daily_rows=daily_rows,
        pulled=pulled,
        ingested=ingested,
        skipped_duplicates=skipped_duplicates,
    )
    _write_sheet(
        workbook,
        title="FilteredEvents",
        columns=FILTERED_COLUMNS,
        rows=filtered_rows,
    )
    _write_sheet(
        workbook,
        title="DailyUsage",
        columns=DAILY_COLUMNS,
        rows=daily_rows,
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path
