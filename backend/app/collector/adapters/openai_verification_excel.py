"""Build Excel verification workbook for OpenAI organization usage/cost pulls."""

from __future__ import annotations

from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from app.collector.adapters.openai_usage_parsing import (
    _parse_cost_amount,
    _result_rows,
    _usage_row_has_activity,
)

VERIFICATION_FILENAME = "calculation-verification.xlsx"

API_COLUMNS = [
    ("source", "Source"),
    ("page", "Page"),
    ("status_code", "HTTP"),
    ("url", "URL"),
    ("bucket_count", "Buckets"),
    ("result_rows", "Result rows"),
]

USAGE_COLUMNS = [
    ("source", "Source"),
    ("bucket_start", "Bucket start"),
    ("model", "Model"),
    ("user_id", "User id"),
    ("input_tokens", "Input tokens"),
    ("output_tokens", "Output tokens"),
    ("requests", "Requests"),
    ("parsed", "Parsed for ingest?"),
]

COST_COLUMNS = [
    ("source", "Source"),
    ("bucket_start", "Bucket start"),
    ("line_item", "Line item"),
    ("project_id", "Project id"),
    ("amount_usd", "Amount (USD)"),
]

FINAL_COLUMNS = [
    ("vendor_event_id", "vendor_event_id"),
    ("occurred_at", "occurred_at"),
    ("model", "model"),
    ("user_email", "user_email"),
    ("input_tokens", "input_tokens"),
    ("output_tokens", "output_tokens"),
    ("total_tokens", "total_tokens"),
    ("estimated_cost", "estimated_cost (USD)"),
    ("requests", "requests"),
]


def _sheet_from_rows(
    workbook: Workbook,
    title: str,
    columns: list[tuple[str, str]],
    rows: list[dict[str, Any]],
) -> None:
    sheet = workbook.create_sheet(title=title)
    headers = [label for _, label in columns]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    keys = [key for key, _ in columns]
    for row in rows:
        sheet.append([row.get(key) for key in keys])
    for index, (_, label) in enumerate(columns, start=1):
        width = min(max(len(label), 12), 48)
        sheet.column_dimensions[get_column_letter(index)].width = width


def build_verification_rows(
    raw_pages: list[dict[str, Any]],
    parsed_records: list[dict[str, Any]] | None = None,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    api_rows: list[dict[str, Any]] = []
    usage_rows: list[dict[str, Any]] = []
    cost_rows: list[dict[str, Any]] = []

    for entry in raw_pages:
        source = str(entry.get("source") or "")
        page = int(entry.get("page") or 0)
        status_code = entry.get("status_code")
        url = str(entry.get("url") or "")
        payload = entry.get("response")
        buckets = payload.get("data", []) if isinstance(payload, dict) else []
        result_count = 0
        if isinstance(buckets, list):
            for bucket in buckets:
                if not isinstance(bucket, dict):
                    continue
                rows = _result_rows(bucket)
                result_count += len(rows)
                bucket_start = bucket.get("start_time_iso") or bucket.get("start_time")
                if "cost" in source:
                    for row in rows:
                        cost_rows.append(
                            {
                                "source": source,
                                "bucket_start": bucket_start,
                                "line_item": row.get("line_item"),
                                "project_id": row.get("project_id"),
                                "amount_usd": float(_parse_cost_amount(row)),
                            }
                        )
                else:
                    for row in rows:
                        usage_rows.append(
                            {
                                "source": source,
                                "bucket_start": bucket_start,
                                "model": row.get("model"),
                                "user_id": row.get("user_id"),
                                "input_tokens": row.get("input_tokens"),
                                "output_tokens": row.get("output_tokens"),
                                "requests": row.get("num_model_requests") or row.get("n_requests"),
                                "parsed": "Y" if _usage_row_has_activity(row) else "N",
                            }
                        )
        api_rows.append(
            {
                "source": source,
                "page": page,
                "status_code": status_code,
                "url": url,
                "bucket_count": len(buckets) if isinstance(buckets, list) else 0,
                "result_rows": result_count,
            }
        )

    final_rows: list[dict[str, Any]] = []
    for record in parsed_records or []:
        input_tokens = int(record.get("input_tokens") or 0)
        output_tokens = int(record.get("output_tokens") or 0)
        cache_read = int(record.get("cache_read_tokens") or 0)
        cache_write = int(record.get("cache_write_tokens") or 0)
        try:
            estimated_cost = float(Decimal(str(record.get("estimated_cost") or 0)))
        except Exception:  # noqa: BLE001
            estimated_cost = 0.0
        final_rows.append(
            {
                "vendor_event_id": record.get("vendor_event_id"),
                "occurred_at": record.get("occurred_at"),
                "model": record.get("model"),
                "user_email": record.get("user_email"),
                "input_tokens": input_tokens,
                "output_tokens": output_tokens,
                "total_tokens": input_tokens + output_tokens + cache_read + cache_write,
                "estimated_cost": estimated_cost,
                "requests": record.get("requests"),
            }
        )

    return api_rows, usage_rows, cost_rows, final_rows


def write_calculation_verification_excel(
    *,
    output_path: Path,
    raw_pages: list[dict[str, Any]],
    parsed_records: list[dict[str, Any]] | None = None,
    pulled: int = 0,
    ingested: int = 0,
    skipped_duplicates: int = 0,
    **_kwargs: Any,
) -> Path:
    api_rows, usage_rows, cost_rows, final_rows = build_verification_rows(
        raw_pages,
        parsed_records=parsed_records,
    )
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    summary.append(["Metric", "Value"])
    summary["A1"].font = Font(bold=True)
    summary["B1"].font = Font(bold=True)
    summary.append(["Pulled records", pulled])
    summary.append(["Ingested new", ingested])
    summary.append(["Skipped duplicates", skipped_duplicates])
    summary.append(["Raw API pages", len(raw_pages)])
    summary.append(["Usage result rows", len(usage_rows)])
    summary.append(["Cost result rows", len(cost_rows)])
    summary.append(["Final parsed records", len(final_rows)])
    total_tokens = sum(int(row.get("total_tokens") or 0) for row in final_rows)
    total_cost = sum(float(row.get("estimated_cost") or 0) for row in final_rows)
    summary.append(["Final total tokens", total_tokens])
    summary.append(["Final total cost (USD)", round(total_cost, 6)])

    _sheet_from_rows(workbook, "ApiCalls", API_COLUMNS, api_rows)
    _sheet_from_rows(workbook, "UsagePull", USAGE_COLUMNS, usage_rows)
    _sheet_from_rows(workbook, "CostsPull", COST_COLUMNS, cost_rows)
    _sheet_from_rows(workbook, "FinalRecords", FINAL_COLUMNS, final_rows)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path
