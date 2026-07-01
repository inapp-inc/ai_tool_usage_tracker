"""Read GitHub Copilot billing Excel exports for CSV-compatible parsing."""

from __future__ import annotations

import csv
import io
from typing import Any

from openpyxl import load_workbook


def xlsx_to_csv_text(content: bytes) -> tuple[str, list[str], str | None]:
    """Convert the first worksheet of an Excel file to CSV text and headers."""
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        return "", [], f"Invalid Excel file: {exc}"

    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return "", [], "Excel file is empty."

    header_cells = rows[0]
    headers = [_cell_text(value) for value in header_cells]
    headers = [header for header in headers if header]
    if not headers:
        return "", [], "Excel file is missing a header row."

    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(headers)
    for row in rows[1:]:
        if row is None or all(_is_blank(value) for value in row):
            continue
        writer.writerow([_cell_text(value) for value in row[: len(headers)]])

    return buffer.getvalue(), headers, None


def _is_blank(value: Any) -> bool:
    return value is None or str(value).strip() == ""


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if hasattr(value, "isoformat"):
        try:
            return value.isoformat()[:10]
        except Exception:  # noqa: BLE001
            pass
    return str(value).strip()
